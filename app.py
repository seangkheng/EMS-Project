# app.py (Final Corrected Version with All Features & PDF Image Fix for Render)

import os
import time
import sqlite3
import io
import pandas as pd
import calendar
import traceback
import math
import base64 # Import for image encoding
import mimetypes # Import for getting image type
from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
from flask_bcrypt import Bcrypt
import jwt
from datetime import datetime, timedelta, timezone
from werkzeug.utils import secure_filename
from functools import wraps
from weasyprint import HTML, CSS
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

load_dotenv()

# --- App Initialization & Configs ---
app = Flask(__name__, static_folder='static', static_url_path='/static')
CORS(app)
bcrypt = Bcrypt(app)

# --- Configuration for Render Deployment ---
# On Render, a persistent disk is mounted at '/var/data'.
# We'll store the database and uploads there to prevent data loss on deploys.
IS_ON_RENDER = os.environ.get('RENDER', False)
DATA_DIR = '/var/data' if IS_ON_RENDER else os.path.dirname(os.path.abspath(__file__))

DATABASE_FILE = os.path.join(DATA_DIR, 'ems_database.db')
UPLOAD_FOLDER = os.path.join(DATA_DIR, 'uploads')

app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'a_default_secret_key_if_not_set_for_dev')
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FONT_FOLDER = os.path.join(BASE_DIR, 'fonts')
KHMER_TTF = os.path.join(FONT_FOLDER, 'KhmerOS.ttf')
JAPANESE_TTF = os.path.join(FONT_FOLDER, 'NotoSansJP-Regular.ttf')

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Create necessary directories if they don't exist
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
    print(f"--- INFO: Created uploads directory at {UPLOAD_FOLDER} ---")

if not os.path.exists(FONT_FOLDER):
    # This should ideally be part of your git repo, but we create it just in case.
    os.makedirs(FONT_FOLDER)
    print(f"--- WARNING: Fonts directory not found at {FONT_FOLDER}. PDF exports might fail. ---")


# --- Database Helper Function ---
def get_db_connection():
    conn = sqlite3.connect(DATABASE_FILE)
    conn.row_factory = sqlite3.Row
    return conn

# --- INITIAL DATABASE AND ADMIN SETUP ---
def setup_database_and_admin():
    print("--- INFO: Checking database and setting up default admin... ---")
    print(f"--- INFO: Database file path: {DATABASE_FILE} ---")
    conn = get_db_connection()
    cursor = conn.cursor()
    sql_commands = [
        """CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT NOT NULL UNIQUE, password TEXT NOT NULL,
            email TEXT NOT NULL UNIQUE, full_name TEXT, role TEXT NOT NULL DEFAULT 'teacher',
            is_active INTEGER NOT NULL DEFAULT 1, last_login TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );""",
        """CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, name_km TEXT, name_en TEXT, name_jp TEXT,
            dob TEXT NOT NULL, contact TEXT NOT NULL,
            address TEXT, photo_filename TEXT, parent_name TEXT, parent_contact TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );""",
        """CREATE TABLE IF NOT EXISTS teachers (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, email TEXT NOT NULL UNIQUE, contact TEXT NOT NULL,
            specialty TEXT, hire_date TEXT, photo_filename TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );""",
        """CREATE TABLE IF NOT EXISTS subjects (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL UNIQUE, description TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );""",
        """CREATE TABLE IF NOT EXISTS classes (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, teacher_id INTEGER, subject_id INTEGER,
            academic_year TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (teacher_id) REFERENCES teachers(id) ON DELETE SET NULL,
            FOREIGN KEY (subject_id) REFERENCES subjects(id) ON DELETE SET NULL
        );""",
        """CREATE TABLE IF NOT EXISTS enrollments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL, class_id INTEGER NOT NULL,
            enrollment_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE CASCADE,
            FOREIGN KEY (class_id) REFERENCES classes(id) ON DELETE CASCADE,
            UNIQUE(student_id, class_id)
        );""",
        """CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL, class_id INTEGER NOT NULL,
            attendance_date TEXT NOT NULL, status TEXT NOT NULL,
            FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE CASCADE,
            FOREIGN KEY (class_id) REFERENCES classes(id) ON DELETE CASCADE,
            UNIQUE(student_id, class_id, attendance_date)
        );""",
        """CREATE TABLE IF NOT EXISTS grades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            class_id INTEGER NOT NULL,
            subject_id INTEGER NOT NULL,
            exam_type TEXT NOT NULL, -- e.g., 'Monthly', 'Final'
            score REAL, -- Using REAL for potential decimal scores
            grade_date TEXT NOT NULL,
            FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE CASCADE,
            FOREIGN KEY (class_id) REFERENCES classes(id) ON DELETE CASCADE,
            FOREIGN KEY (subject_id) REFERENCES subjects(id) ON DELETE CASCADE,
            UNIQUE(student_id, class_id, subject_id, exam_type, grade_date)
        );""",
        """CREATE TABLE IF NOT EXISTS timetables (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_id INTEGER NOT NULL,
            teacher_id INTEGER NOT NULL,
            subject_id INTEGER NOT NULL,
            day_of_week INTEGER NOT NULL, -- 1=ចន្ទ, 2=អង្គារ, ... 7=អាទិត្យ
            start_time TEXT NOT NULL, -- ឧទាហរណ៍: '08:00'
            end_time TEXT NOT NULL,   -- ឧទាហรณ์: '09:00'
            FOREIGN KEY (class_id) REFERENCES classes(id) ON DELETE CASCADE,
            FOREIGN KEY (teacher_id) REFERENCES teachers(id) ON DELETE CASCADE,
            FOREIGN KEY (subject_id) REFERENCES subjects(id) ON DELETE CASCADE
        );""",
        """CREATE TABLE IF NOT EXISTS announcements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            content TEXT NOT NULL,
            user_id INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        );"""
    ]
    for command in sql_commands:
        try:
            cursor.execute(command)
        except sqlite3.OperationalError as e:
            if 'duplicate column name' not in str(e): print(f"--- WARNING: Harmless error during table creation: {e}")

    try: cursor.execute("ALTER TABLE students ADD COLUMN parent_name TEXT;")
    except sqlite3.OperationalError: pass
    try: cursor.execute("ALTER TABLE students ADD COLUMN parent_contact TEXT;")
    except sqlite3.OperationalError: pass
    try: cursor.execute("ALTER TABLE students ADD COLUMN name_km TEXT;")
    except sqlite3.OperationalError: pass
    try: cursor.execute("ALTER TABLE students ADD COLUMN name_en TEXT;")
    except sqlite3.OperationalError: pass
    try: cursor.execute("ALTER TABLE students ADD COLUMN name_jp TEXT;")
    except sqlite3.OperationalError: pass


    default_password = "admin123"
    hashed_password = bcrypt.generate_password_hash(default_password).decode('utf-8')
    cursor.execute("SELECT * FROM users WHERE username = 'admin'")
    user = cursor.fetchone()
    if user:
        cursor.execute("UPDATE users SET password = ?, role = 'admin', is_active = 1, email = ? WHERE username = 'admin'", [hashed_password, 'seangkhenghou@ibmscam.work'])
    else:
        cursor.execute("INSERT INTO users(username, password, email, full_name, role, is_active) VALUES (?, ?, ?, ?, ?, ?)",
            ('admin', hashed_password, 'seangkhenghou@ibmscam.work', 'Default Admin', 'admin', 1))
    conn.commit()
    conn.close()
    print(f"--- INFO: Setup complete. You can log in with Username: admin, Password: {default_password} ---")


def get_token_data():
    token = None
    if 'authorization' in request.headers:
        try:
            token = request.headers['authorization'].split(' ')[1]
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=["HS256"])
            return data
        except Exception: return None
    return None

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        data = get_token_data()
        if data is None: return jsonify({'message': 'Token is invalid or missing!'}), 401
        kwargs['current_user'] = data
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        data = get_token_data()
        if data is None or data.get('role') != 'admin': return jsonify({'message': 'Admin access required!'}), 403
        kwargs['current_user'] = data
        return f(*args, **kwargs)
    return decorated

# --- Main Route to Serve Frontend ---
@app.route('/')
def serve_index():
    return send_from_directory('.', 'index.html')

# --- Route to serve uploaded files from the persistent disk ---
@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/<path:path>')
def serve_static_or_index(path):
    # This will handle static files like CSS, JS, and images from the 'static' folder
    # as well as routing for the single-page application.
    if not os.path.splitext(path)[1] and path != 'favicon.ico':
        return send_from_directory('.', 'index.html')
    return send_from_directory('.', path)


# --- API Routes ---
@app.route('/api/login', methods=['POST'])
def login():
    try:
        data = request.get_json()
        username, password = data.get('username'), data.get('password')
        conn = get_db_connection()
        user = conn.execute("SELECT * FROM users WHERE username = ?", [username]).fetchone()
        if user and user['password'] and bcrypt.check_password_hash(user['password'], password):
            if not user['is_active']:
                conn.close()
                return jsonify({'message': 'This account has been deactivated.'}), 403
            conn.execute("UPDATE users SET last_login = ? WHERE id = ?", (datetime.now(timezone.utc).isoformat(), user['id']))
            conn.commit()
            conn.close()
            token = jwt.encode({'id': user['id'], 'username': user['username'], 'role': user['role'], 'exp': datetime.now(timezone.utc) + timedelta(hours=24)}, app.config['SECRET_KEY'], algorithm="HS256")
            return jsonify({'message': 'Login successful!', 'token': token})
        conn.close()
        return jsonify({'message': 'Invalid username or password'}), 401
    except Exception as e:
        print(f"Login Error: {e}")
        return jsonify({'error': 'An internal server error occurred'}), 500

# --- Dashboard API ---
@app.route('/api/dashboard/stats', methods=['GET'])
@token_required
def get_dashboard_stats(**kwargs):
    conn = get_db_connection()
    students = conn.execute("SELECT COUNT(id) as total FROM students").fetchone()['total']
    teachers = conn.execute("SELECT COUNT(id) as total FROM teachers").fetchone()['total']
    classes = conn.execute("SELECT COUNT(id) as total FROM classes").fetchone()['total']
    users = conn.execute("SELECT COUNT(id) as total FROM users WHERE is_active = 1").fetchone()['total']
    conn.close()
    return jsonify({'students': students, 'teachers': teachers, 'classes': classes, 'active_users': users})

@app.route('/api/dashboard/class-sizes', methods=['GET'])
@token_required
def get_class_sizes(**kwargs):
    conn = get_db_connection()
    class_sizes = conn.execute("""
        SELECT c.name as class_name, COUNT(e.student_id) as student_count
        FROM classes c
        LEFT JOIN enrollments e ON c.id = e.class_id
        GROUP BY c.id, c.name
        ORDER BY student_count DESC
    """).fetchall()
    conn.close()
    return jsonify([dict(row) for row in class_sizes])

@app.route('/api/dashboard/user-roles', methods=['GET'])
@token_required
def get_user_roles(**kwargs):
    conn = get_db_connection()
    user_roles = conn.execute("""
        SELECT role, COUNT(id) as count
        FROM users
        WHERE is_active = 1
        GROUP BY role
    """).fetchall()
    conn.close()
    return jsonify([dict(row) for row in user_roles])


# == User API ==
@app.route('/api/users', methods=['GET'])
@admin_required
def get_users(**kwargs):
    page = request.args.get('page', 1, type=int)
    per_page = 15
    offset = (page - 1) * per_page
    
    conn = get_db_connection()

    total_items = conn.execute("SELECT COUNT(id) as total FROM users").fetchone()['total']
    total_pages = math.ceil(total_items / per_page)

    users = conn.execute("SELECT id, username, email, full_name, role, is_active, last_login FROM users ORDER BY id DESC LIMIT ? OFFSET ?", (per_page, offset)).fetchall()
    conn.close()
    return jsonify({
        'data': [dict(row) for row in users],
        'current_page': page,
        'total_pages': total_pages,
        'total_items': total_items
    })

@app.route('/api/register', methods=['POST'])
@admin_required
def register(**kwargs):
    data = request.get_json()
    username, password, email, full_name, role = data.get('username'), data.get('password'), data.get('email'), data.get('full_name'), data.get('role', 'teacher')
    if not all([username, password, email, full_name, role]):
        return jsonify({'message': 'All fields are required.'}), 400
    hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
    conn = get_db_connection()
    try:
        conn.execute("INSERT INTO users(username, password, email, full_name, role) VALUES (?, ?, ?, ?, ?)", (username, hashed_password, email, full_name, role))
        conn.commit()
        return jsonify({'message': 'User registered successfully!'}), 201
    except sqlite3.Error as e:
        conn.rollback()
        print(f"DATABASE ERROR in register: {e}")
        return jsonify({'message': f'Database error: {e}'}), 500
    finally:
        if conn:
            conn.close()

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@admin_required
def update_user(user_id, **kwargs):
    data = request.get_json()
    role, is_active = data.get('role'), data.get('is_active')
    if role is None and is_active is None:
        return jsonify({'message': 'No role or status provided to update'}), 400
    conn = get_db_connection()
    try:
        if role and is_active is not None:
            conn.execute("UPDATE users SET role = ?, is_active = ? WHERE id = ?", (role, is_active, user_id))
        elif role:
            conn.execute("UPDATE users SET role = ? WHERE id = ?", (role, user_id))
        elif is_active is not None:
            conn.execute("UPDATE users SET is_active = ? WHERE id = ?", (is_active, user_id))
        conn.commit()
        return jsonify({'message': 'User updated successfully!'})
    except sqlite3.Error as e:
        conn.rollback()
        print(f"DATABASE ERROR in update_user: {e}")
        return jsonify({'message': f'Database error: {e}'}), 500
    finally:
        if conn:
            conn.close()

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def delete_user(user_id, **kwargs):
    if user_id == kwargs['current_user']['id']:
        return jsonify({'message': 'Admin cannot delete themselves'}), 403
    conn = get_db_connection()
    conn.execute("DELETE FROM users WHERE id = ?", [user_id])
    conn.commit()
    conn.close()
    return jsonify({'message': 'User deleted successfully!'})

# == Student API ==
@app.route('/api/students', methods=['GET'])
@token_required
def get_students(**kwargs):
    page = request.args.get('page', 1, type=int)
    per_page = 15
    offset = (page - 1) * per_page
    search_term = request.args.get('search', '')

    conn = get_db_connection()
    
    count_query = "SELECT COUNT(DISTINCT s.id) as total FROM students s"
    base_query = """
        SELECT s.*, c.name as class_name, e.class_id
        FROM students s
        LEFT JOIN enrollments e ON s.id = e.student_id
        LEFT JOIN classes c ON e.class_id = c.id
    """
    params = []
    
    if search_term:
        where_clause = " WHERE s.name_km LIKE ? OR s.name_en LIKE ? OR s.name_jp LIKE ?"
        count_query += where_clause
        base_query += where_clause
        params.extend([f'%{search_term}%', f'%{search_term}%', f'%{search_term}%'])

    total_items = conn.execute(count_query, params).fetchone()['total']
    total_pages = math.ceil(total_items / per_page)

    base_query += " GROUP BY s.id ORDER BY s.id DESC LIMIT ? OFFSET ?"
    params.extend([per_page, offset])
    
    students = conn.execute(base_query, params).fetchall()
    conn.close()

    return jsonify({
        'data': [dict(row) for row in students],
        'current_page': page,
        'total_pages': total_pages,
        'total_items': total_items
    })

@app.route('/api/students/<int:id>', methods=['GET'])
@token_required
def get_student_details(id, **kwargs):
    conn = get_db_connection()
    student = conn.execute("""
        SELECT s.*, c.name as class_name
        FROM students s
        LEFT JOIN enrollments e ON s.id = e.student_id
        LEFT JOIN classes c ON e.class_id = c.id
        WHERE s.id = ?
        GROUP BY s.id
    """, (id,)).fetchone()

    if student is None:
        return jsonify({'message': 'Student not found'}), 404

    attendance = conn.execute("""
        SELECT attendance_date, status FROM attendance
        WHERE student_id = ? ORDER BY attendance_date DESC
    """, (id,)).fetchall()

    conn.close()

    student_details = dict(student)
    student_details['attendance_history'] = [dict(row) for row in attendance]

    return jsonify(student_details)


@app.route('/api/students', methods=['POST'])
@admin_required
def add_student(**kwargs):
    conn = get_db_connection()
    try:
        form_data = request.form
        name_km = form_data.get('name_km')
        name_en = form_data.get('name_en')
        name_jp = form_data.get('name_jp')
        dob = form_data.get('dob')
        contact = form_data.get('contact')
        address = form_data.get('address')
        parent_name = form_data.get('parent_name')
        parent_contact = form_data.get('parent_contact')
        class_id = form_data.get('class_id')
        photo = request.files.get('photo')

        photo_filename = None
        if photo and photo.filename != '':
             photo_filename = secure_filename(str(int(time.time())) + os.path.splitext(photo.filename)[1])
             photo.save(os.path.join(app.config['UPLOAD_FOLDER'], photo_filename))


        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO students (name, name_km, name_en, name_jp, dob, contact, address, parent_name, parent_contact, photo_filename)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (name_km, name_km, name_en, name_jp, dob, contact, address, parent_name, parent_contact, photo_filename))

        if class_id:
            student_id = cursor.lastrowid
            cursor.execute("INSERT INTO enrollments (student_id, class_id) VALUES (?, ?)", (student_id, class_id))

        conn.commit()
        return jsonify({'message': 'Student added successfully!'}), 201
    except sqlite3.Error as e:
        conn.rollback()
        print(f"DATABASE ERROR in add_student: {e}")
        return jsonify({'message': f'Database error: {e}'}), 500
    finally:
        if conn:
            conn.close()

@app.route('/api/students/<int:id>', methods=['PUT'])
@admin_required
def update_student(id, **kwargs):
    conn = get_db_connection()
    try:
        form_data = request.form
        name_km = form_data.get('name_km')
        name_en = form_data.get('name_en')
        name_jp = form_data.get('name_jp')
        dob = form_data.get('dob')
        contact = form_data.get('contact')
        address = form_data.get('address')
        parent_name = form_data.get('parent_name')
        parent_contact = form_data.get('parent_contact')
        class_id = form_data.get('class_id')
        photo = request.files.get('photo')

        update_query = """
            UPDATE students SET
            name=?, name_km=?, name_en=?, name_jp=?, dob=?, contact=?, address=?, parent_name=?, parent_contact=?
            WHERE id=?
        """
        params = (name_km, name_km, name_en, name_jp, dob, contact, address, parent_name, parent_contact, id)

        if photo and photo.filename != '':
            old_photo_row = conn.execute("SELECT photo_filename FROM students WHERE id = ?", [id]).fetchone()
            if old_photo_row and old_photo_row['photo_filename']:
                old_photo_path = os.path.join(app.config['UPLOAD_FOLDER'], old_photo_row['photo_filename'])
                if os.path.exists(old_photo_path):
                    os.remove(old_photo_path)
            
            photo_filename = secure_filename(str(int(time.time())) + os.path.splitext(photo.filename)[1])
            photo.save(os.path.join(app.config['UPLOAD_FOLDER'], photo_filename))

            update_query = """
                UPDATE students SET
                name=?, name_km=?, name_en=?, name_jp=?, dob=?, contact=?, address=?, parent_name=?, parent_contact=?, photo_filename=?
                WHERE id=?
            """
            params = (name_km, name_km, name_en, name_jp, dob, contact, address, parent_name, parent_contact, photo_filename, id)

        conn.execute(update_query, params)

        # Update class enrollment
        conn.execute("DELETE FROM enrollments WHERE student_id = ?", [id])
        if class_id:
            conn.execute("INSERT INTO enrollments (student_id, class_id) VALUES (?, ?)", (id, class_id))

        conn.commit()
        return jsonify({'message': 'Student updated successfully!'})
    except sqlite3.Error as e:
        conn.rollback()
        print(f"DATABASE ERROR in update_student: {e}")
        return jsonify({'message': f'Database error: {e}'}), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/students/<int:id>', methods=['DELETE'])
@admin_required
def delete_student(id, **kwargs):
    conn = get_db_connection()
    photo_to_delete = conn.execute("SELECT photo_filename FROM students WHERE id = ?", [id]).fetchone()
    conn.execute("DELETE FROM students WHERE id = ?", [id])
    conn.commit()
    conn.close()
    if photo_to_delete and photo_to_delete['photo_filename']:
        photo_path = os.path.join(app.config['UPLOAD_FOLDER'], photo_to_delete['photo_filename'])
        if os.path.exists(photo_path):
            os.remove(photo_path)
    return jsonify({'message': 'Student deleted successfully!'})

# == Teacher API ==
@app.route('/api/teachers', methods=['GET'])
@token_required
def get_teachers(**kwargs):
    page = request.args.get('page', 1, type=int)
    per_page = 15
    offset = (page - 1) * per_page
    search_term = request.args.get('search', '')

    conn = get_db_connection()
    
    count_query = "SELECT COUNT(id) as total FROM teachers"
    base_query = "SELECT * FROM teachers"
    params = []
    
    if search_term:
        where_clause = " WHERE name LIKE ? OR email LIKE ?"
        count_query += where_clause
        base_query += where_clause
        params.extend([f'%{search_term}%', f'%{search_term}%'])

    total_items = conn.execute(count_query, params).fetchone()['total']
    total_pages = math.ceil(total_items / per_page)
    
    base_query += " ORDER BY id DESC LIMIT ? OFFSET ?"
    params.extend([per_page, offset])
    
    teachers = conn.execute(base_query, params).fetchall()
    conn.close()
    
    return jsonify({
        'data': [dict(row) for row in teachers],
        'current_page': page,
        'total_pages': total_pages,
        'total_items': total_items
    })

@app.route('/api/teachers/<int:id>', methods=['GET'])
@token_required
def get_teacher_details(id, **kwargs):
    conn = get_db_connection()
    teacher = conn.execute("SELECT * FROM teachers WHERE id = ?", (id,)).fetchone()

    if teacher is None:
        return jsonify({'message': 'Teacher not found'}), 404

    classes = conn.execute("""
        SELECT c.id, c.name, c.academic_year, s.name as subject_name
        FROM classes c
        LEFT JOIN subjects s ON c.subject_id = s.id
        WHERE c.teacher_id = ?
    """, (id,)).fetchall()

    conn.close()

    teacher_details = dict(teacher)
    teacher_details['assigned_classes'] = [dict(row) for row in classes]

    return jsonify(teacher_details)


@app.route('/api/teachers', methods=['POST'])
@admin_required
def add_teacher(**kwargs):
    conn = get_db_connection()
    try:
        name, email, contact, specialty, hire_date = request.form.get('name'), request.form.get('email'), request.form.get('contact'), request.form.get('specialty'), request.form.get('hire_date')
        photo = request.files.get('photo')
        photo_filename = None
        if photo and photo.filename != '':
             photo_filename = secure_filename(str(int(time.time())) + os.path.splitext(photo.filename)[1])
             photo.save(os.path.join(app.config['UPLOAD_FOLDER'], photo_filename))

        conn.execute("INSERT INTO teachers (name, email, contact, specialty, hire_date, photo_filename) VALUES (?, ?, ?, ?, ?, ?)", (name, email, contact, specialty, hire_date, photo_filename))
        conn.commit()
        return jsonify({'message': 'Teacher added successfully!'}), 201
    except sqlite3.Error as e:
        conn.rollback()
        print(f"DATABASE ERROR in add_teacher: {e}")
        return jsonify({'message': f'Database error: {e}'}), 500
    finally:
        if conn:
            conn.close()

@app.route('/api/teachers/<int:id>', methods=['PUT'])
@admin_required
def update_teacher(id, **kwargs):
    conn = get_db_connection()
    try:
        name, email, contact, specialty, hire_date = request.form.get('name'), request.form.get('email'), request.form.get('contact'), request.form.get('specialty'), request.form.get('hire_date')
        photo = request.files.get('photo')
        if photo and photo.filename != '':
            old_photo_row = conn.execute("SELECT photo_filename FROM teachers WHERE id = ?", [id]).fetchone()
            if old_photo_row and old_photo_row['photo_filename']:
                old_photo_path = os.path.join(app.config['UPLOAD_FOLDER'], old_photo_row['photo_filename'])
                if os.path.exists(old_photo_path):
                    os.remove(old_photo_path)
            
            photo_filename = secure_filename(str(int(time.time())) + os.path.splitext(photo.filename)[1])
            photo.save(os.path.join(app.config['UPLOAD_FOLDER'], photo_filename))
            conn.execute("UPDATE teachers SET name=?, email=?, contact=?, specialty=?, hire_date=?, photo_filename=? WHERE id=?", (name, email, contact, specialty, hire_date, photo_filename, id))
        else:
            conn.execute("UPDATE teachers SET name=?, email=?, contact=?, specialty=?, hire_date=? WHERE id=?", (name, email, contact, specialty, hire_date, id))
        conn.commit()
        return jsonify({'message': 'Teacher updated successfully!'})
    except sqlite3.Error as e:
        conn.rollback()
        print(f"DATABASE ERROR in update_teacher: {e}")
        return jsonify({'message': f'Database error: {e}'}), 500
    finally:
        if conn:
            conn.close()

@app.route('/api/teachers/<int:id>', methods=['DELETE'])
@admin_required
def delete_teacher(id, **kwargs):
    conn = get_db_connection()
    photo_to_delete = conn.execute("SELECT photo_filename FROM teachers WHERE id = ?", [id]).fetchone()
    conn.execute("DELETE FROM teachers WHERE id = ?", [id])
    conn.commit()
    conn.close()
    if photo_to_delete and photo_to_delete['photo_filename']:
        photo_path = os.path.join(app.config['UPLOAD_FOLDER'], photo_to_delete['photo_filename'])
        if os.path.exists(photo_path):
            os.remove(photo_path)
    return jsonify({'message': 'Teacher deleted successfully!'})

# == Subject API ==
@app.route('/api/subjects', methods=['GET'])
@token_required
def get_subjects(**kwargs):
    page = request.args.get('page', 1, type=int)
    per_page = 15
    offset = (page - 1) * per_page
    
    conn = get_db_connection()

    total_items = conn.execute("SELECT COUNT(id) as total FROM subjects").fetchone()['total']
    total_pages = math.ceil(total_items / per_page)

    subjects = conn.execute("SELECT * FROM subjects ORDER BY id DESC LIMIT ? OFFSET ?", (per_page, offset)).fetchall()
    conn.close()
    return jsonify({
        'data': [dict(row) for row in subjects],
        'current_page': page,
        'total_pages': total_pages,
        'total_items': total_items
    })

@app.route('/api/subjects', methods=['POST'])
@admin_required
def add_subject(**kwargs):
    data = request.get_json()
    name, description = data.get('name'), data.get('description')
    conn = get_db_connection()
    conn.execute("INSERT INTO subjects (name, description) VALUES (?, ?)", (name, description))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Subject added successfully!'}), 201

@app.route('/api/subjects/<int:id>', methods=['PUT'])
@admin_required
def update_subject(id, **kwargs):
    data = request.get_json()
    name, description = data.get('name'), data.get('description')
    conn = get_db_connection()
    conn.execute("UPDATE subjects SET name = ?, description = ? WHERE id = ?", (name, description, id))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Subject updated successfully!'})

@app.route('/api/subjects/<int:id>', methods=['DELETE'])
@admin_required
def delete_subject(id, **kwargs):
    conn = get_db_connection()
    conn.execute("DELETE FROM subjects WHERE id = ?", [id])
    conn.commit()
    conn.close()
    return jsonify({'message': 'Subject deleted successfully!'})

# == Class API ==
@app.route('/api/classes', methods=['GET'])
@token_required
def get_classes(current_user, **kwargs):
    page = request.args.get('page', 1, type=int)
    per_page = 15
    offset = (page - 1) * per_page
    search_term = request.args.get('search', '')
    
    conn = get_db_connection()
    
    count_query = "SELECT COUNT(c.id) as total FROM classes c"
    base_query = """
        SELECT c.*, t.name as teacher_name, s.name as subject_name
        FROM classes c
        LEFT JOIN teachers t ON c.teacher_id = t.id
        LEFT JOIN subjects s ON c.subject_id = s.id
    """
    params = []
    where_conditions = []

    if current_user['role'] == 'teacher':
        user = conn.execute("SELECT email FROM users WHERE id = ?", (current_user['id'],)).fetchone()
        if user:
            teacher = conn.execute("SELECT id FROM teachers WHERE email = ?", (user['email'],)).fetchone()
            if teacher:
                where_conditions.append("c.teacher_id = ?")
                params.append(teacher['id'])
            else:
                return jsonify({'data': [], 'current_page': 1, 'total_pages': 0, 'total_items': 0})
        else:
             return jsonify({'data': [], 'current_page': 1, 'total_pages': 0, 'total_items': 0})

    if search_term:
        where_conditions.append("(c.name LIKE ? OR c.academic_year LIKE ?)")
        params.extend([f'%{search_term}%', f'%{search_term}%'])

    if where_conditions:
        where_clause = " WHERE " + " AND ".join(where_conditions)
        count_query += where_clause
        base_query += where_clause

    total_items = conn.execute(count_query, params).fetchone()['total']
    total_pages = math.ceil(total_items / per_page)

    base_query += " ORDER BY c.id DESC LIMIT ? OFFSET ?"
    params.extend([per_page, offset])
    
    classes = conn.execute(base_query, params).fetchall()
    conn.close()
    
    return jsonify({
        'data': [dict(row) for row in classes],
        'current_page': page,
        'total_pages': total_pages,
        'total_items': total_items
    })

@app.route('/api/classes', methods=['POST'])
@admin_required
def add_class(**kwargs):
    data = request.get_json()
    name, teacher_id, subject_id, academic_year = data.get('name'), data.get('teacher_id'), data.get('subject_id'), data.get('academic_year')
    conn = get_db_connection()
    conn.execute("INSERT INTO classes (name, teacher_id, subject_id, academic_year) VALUES (?, ?, ?, ?)", (name, teacher_id, subject_id, academic_year))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Class created successfully!'}), 201

@app.route('/api/classes/<int:id>', methods=['PUT'])
@admin_required
def update_class(id, **kwargs):
    data = request.get_json()
    name, teacher_id, subject_id, academic_year = data.get('name'), data.get('teacher_id'), data.get('subject_id'), data.get('academic_year')
    conn = get_db_connection()
    conn.execute("UPDATE classes SET name = ?, teacher_id = ?, subject_id = ?, academic_year = ? WHERE id = ?", (name, teacher_id, subject_id, academic_year, id))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Class updated successfully!'})

@app.route('/api/classes/<int:id>', methods=['DELETE'])
@admin_required
def delete_class(id, **kwargs):
    conn = get_db_connection()
    conn.execute("DELETE FROM classes WHERE id = ?", [id])
    conn.commit()
    conn.close()
    return jsonify({'message': 'Class deleted successfully!'})

# --- Enrollment API Routes ---
@app.route('/api/classes/<int:class_id>/students', methods=['GET'])
@token_required
def get_enrolled_students(class_id, **kwargs):
    conn = get_db_connection()
    students = conn.execute("""
        SELECT s.id, s.name, s.contact, e.id as enrollment_id FROM students s
        JOIN enrollments e ON s.id = e.student_id
        WHERE e.class_id = ?
    """, (class_id,)).fetchall()
    conn.close()
    return jsonify([dict(row) for row in students])

@app.route('/api/enrollments', methods=['POST'])
@admin_required
def enroll_student(**kwargs):
    data = request.get_json()
    student_id, class_id = data.get('student_id'), data.get('class_id')
    if not student_id or not class_id:
        return jsonify({'message': 'Student ID and Class ID are required.'}), 400
    conn = get_db_connection()
    try:
        conn.execute("INSERT INTO enrollments (student_id, class_id) VALUES (?, ?)", (student_id, class_id))
        conn.commit()
        return jsonify({'message': 'Student enrolled successfully!'}), 201
    except sqlite3.IntegrityError:
        return jsonify({'message': 'Student is already enrolled in this class.'}), 409
    finally:
        conn.close()

@app.route('/api/enrollments/<int:enrollment_id>', methods=['DELETE'])
@admin_required
def unenroll_student(enrollment_id, **kwargs):
    conn = get_db_connection()
    conn.execute("DELETE FROM enrollments WHERE id = ?", [enrollment_id])
    conn.commit()
    conn.close()
    return jsonify({'message': 'Student unenrolled successfully!'})

# --- Attendance API Routes ---
@app.route('/api/classes/<int:class_id>/attendance', methods=['GET'])
@token_required
def get_attendance(class_id, **kwargs):
    attendance_date = request.args.get('date')
    if not attendance_date:
        return jsonify({'message': 'Date parameter is required.'}), 400
    conn = get_db_connection()
    attendance_records = conn.execute("""
        SELECT s.id, s.name, a.status FROM students s
        JOIN enrollments e ON s.id = e.student_id
        LEFT JOIN attendance a ON s.id = a.student_id AND a.class_id = ? AND a.attendance_date = ?
        WHERE e.class_id = ?
    """, (class_id, attendance_date, class_id)).fetchall()
    conn.close()
    return jsonify([dict(row) for row in attendance_records])

@app.route('/api/attendance', methods=['POST'])
@token_required
def save_attendance(**kwargs):
    data = request.get_json()
    if not isinstance(data, dict):
        return jsonify({'error': 'Invalid data format: expected a JSON object.'}), 400

    attendance_date, class_id, records = data.get('date'), data.get('class_id'), data.get('records')
    if not all([attendance_date, class_id, records]):
        return jsonify({'message': 'Date, Class ID, and records are required.'}), 400
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        for record in records:
            cursor.execute("""
                INSERT INTO attendance (student_id, class_id, attendance_date, status)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(student_id, class_id, attendance_date) DO UPDATE SET
                status = excluded.status;
            """, (record.get('student_id'), class_id, attendance_date, record.get('status')))
        conn.commit()
        return jsonify({'message': 'Attendance saved successfully!'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route('/api/attendance/report', methods=['GET'])
@token_required
def get_attendance_report(**kwargs):
    class_id_str = request.args.get('class_id')
    month_str = request.args.get('month')

    if not class_id_str or not month_str:
        return jsonify({'message': 'Class ID and month are required.'}), 400

    conn = None
    try:
        class_id = int(class_id_str)
        year, month = map(int, month_str.split('-'))
        if not (1 <= month <= 12):
            raise ValueError("Month is out of range")

        conn = get_db_connection()

        students = conn.execute("""
            SELECT id, name FROM students
            WHERE id IN (SELECT student_id FROM enrollments WHERE class_id = ?)
        """, (class_id,)).fetchall()

        attendance_records = conn.execute("""
            SELECT student_id, attendance_date, status
            FROM attendance
            WHERE class_id = ? AND strftime('%Y-%m', attendance_date) = ?
        """, (class_id, month_str)).fetchall()

        report_data = []
        attendance_map = {}
        for record in attendance_records:
            student_id = record['student_id']
            if student_id not in attendance_map:
                attendance_map[student_id] = {}

            date_parts = record['attendance_date'].split('-')
            if len(date_parts) == 3:
                day = int(date_parts[2])
                attendance_map[student_id][day] = record['status']

        for student in students:
            student_report = {
                'student_id': student['id'],
                'student_name': student['name'],
                'attendance': attendance_map.get(student['id'], {})
            }
            report_data.append(student_report)

        month_details = {
            'year': year,
            'month': month,
            'num_days': calendar.monthrange(year, month)[1]
        }

        return jsonify({'report_data': report_data, 'month_details': month_details})

    except Exception as e:
        print(f"---!!!! REPORT ERROR !!!! --->: {e}")
        traceback.print_exc()
        return jsonify({'message': f'An internal error occurred: {e}'}), 500
    finally:
        if conn:
            conn.close()

# --- Gradebook API Routes ---
@app.route('/api/grades', methods=['POST'])
@token_required
def save_grades(**kwargs):
    data = request.get_json()
    class_id = data.get('class_id')
    subject_id = data.get('subject_id')
    exam_type = data.get('exam_type')
    grade_date = data.get('grade_date')
    grades = data.get('grades') # Expects a list of {'student_id': x, 'score': y}

    if not all([class_id, subject_id, exam_type, grade_date, grades is not None]):
        return jsonify({'message': 'Missing required fields.'}), 400

    conn = get_db_connection()
    try:
        for grade in grades:
            student_id = grade.get('student_id')
            score = grade.get('score')
            # If score is empty string or None, treat as NULL in DB
            score_to_save = float(score) if score not in [None, ''] else None

            conn.execute("""
                INSERT INTO grades (student_id, class_id, subject_id, exam_type, grade_date, score)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(student_id, class_id, subject_id, exam_type, grade_date) DO UPDATE SET
                score = excluded.score;
            """, (student_id, class_id, subject_id, exam_type, grade_date, score_to_save))
        conn.commit()
        return jsonify({'message': 'Grades saved successfully!'})
    except Exception as e:
        conn.rollback()
        print(f"---!!!! GRADE SAVE ERROR !!!! --->: {e}")
        traceback.print_exc()
        return jsonify({'message': f'An error occurred: {e}'}), 500
    finally:
        if conn:
            conn.close()

@app.route('/api/grades/class-view', methods=['GET'])
@token_required
def get_class_grades(**kwargs):
    class_id = request.args.get('class_id')
    subject_id = request.args.get('subject_id')
    exam_type = request.args.get('exam_type')
    grade_date = request.args.get('grade_date')

    if not all([class_id, subject_id, exam_type, grade_date]):
        return jsonify({'message': 'Missing required query parameters.'}), 400

    conn = get_db_connection()
    try:
        # First, get all students in the class
        students = conn.execute("""
            SELECT s.id, s.name_km, s.name_en, s.name_jp FROM students s
            JOIN enrollments e ON s.id = e.student_id
            WHERE e.class_id = ?
        """, (class_id,)).fetchall()

        # Then, get existing grades for this specific context
        grades = conn.execute("""
            SELECT student_id, score FROM grades
            WHERE class_id = ? AND subject_id = ? AND exam_type = ? AND grade_date = ?
        """, (class_id, subject_id, exam_type, grade_date)).fetchall()

        grades_map = {g['student_id']: g['score'] for g in grades}

        # Combine the lists
        results = []
        for student in students:
            results.append({
                'student_id': student['id'],
                'student_name_km': student['name_km'],
                'student_name_en': student['name_en'],
                'student_name_jp': student['name_jp'],
                'score': grades_map.get(student['id'])
            })

        return jsonify(results)
    except Exception as e:
        print(f"---!!!! GET GRADES ERROR !!!! --->: {e}")
        traceback.print_exc()
        return jsonify({'message': f'An error occurred: {e}'}), 500
    finally:
        if conn:
            conn.close()

@app.route('/api/grades/export/excel', methods=['POST'])
@token_required
def export_grade_sheet(**kwargs):
    try:
        data = request.get_json()
        grades_data = data.get('grades')
        class_name = data.get('className')
        subject_name = data.get('subjectName')
        exam_type = data.get('examType')
        lang = data.get('lang', 'km')

        if not all([grades_data, class_name, subject_name, exam_type]):
            return jsonify({'message': 'Missing required data for export.'}), 400
            
        headers_translations = {
            'km': ['ឈ្មោះ (ខ្មែរ)', 'ឈ្មោះ (អង់គ្លេស)', 'ឈ្មោះ (ជប៉ុន)', 'ពិន្ទុ'],
            'en': ['Name (Khmer)', 'Name (English)', 'Name (Japanese)', 'Score'],
            'jp': ['名前 (クメール語)', '名前 (英語)', '名前 (日本語)', '点数']
        }
        headers = headers_translations.get(lang, headers_translations['km'])

        wb = Workbook()
        ws = wb.active
        ws.title = f"{exam_type} Grades"

        # Add report title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = f"Grade Sheet: {class_name} - {subject_name} ({exam_type})"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        ws.append(headers)

        if lang == 'km':
            font_name = "Khmer OS Battambang"
        elif lang == 'jp':
            font_name = "MS Gothic"
        else:
            font_name = "Arial"
            
        main_font = Font(name=font_name, size=11)
        bold_font = Font(name=font_name, size=12, bold=True)

        for cell in ws[3]: # Headers are on row 3
            cell.font = bold_font

        for row_data in grades_data:
            ws.append([
                row_data.get('student_name_km', ''),
                row_data.get('student_name_en', ''),
                row_data.get('student_name_jp', ''),
                row_data.get('score', '')
            ])

        for row_cells in ws.iter_rows(min_row=4):
            for cell in row_cells:
                cell.font = main_font
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name=f"grade_sheet_{class_name}.xlsx", as_attachment=True)

    except Exception as e:
        print(f"---!!!! GRADE EXPORT ERROR !!!! --->: {e}")
        traceback.print_exc()
        return jsonify({'message': f'An error occurred during Excel export: {e}'}), 500


# --- Results API Route ---
@app.route('/api/results/class-report', methods=['GET'])
@token_required
def get_class_results(**kwargs):
    class_id_str = request.args.get('class_id')
    exam_type = request.args.get('exam_type')

    if not class_id_str or not exam_type:
        return jsonify({'message': 'Class ID and Exam Type are required.'}), 400

    conn = None
    try:
        class_id = int(class_id_str)
        conn = get_db_connection()

        # 1. Get Class and Teacher Info
        class_info = conn.execute("""
            SELECT c.name as class_name, c.academic_year, t.name as teacher_name
            FROM classes c
            LEFT JOIN teachers t ON c.teacher_id = t.id
            WHERE c.id = ?
        """, (class_id,)).fetchone()

        if not class_info:
            return jsonify({'message': 'Class not found.'}), 404

        # 2. Get all students in the class
        students = conn.execute("""
            SELECT id, name FROM students
            WHERE id IN (SELECT student_id FROM enrollments WHERE class_id = ?)
        """, (class_id,)).fetchall()

        # 3. Get all grades for these students for the given exam type
        grades = conn.execute("""
            SELECT g.student_id, g.score, s.name as subject_name
            FROM grades g
            JOIN subjects s ON g.subject_id = s.id
            WHERE g.class_id = ? AND g.exam_type = ?
        """, (class_id, exam_type)).fetchall()

        # 4. Get attendance summary for these students
        attendance_summary = conn.execute("""
            SELECT student_id, status, COUNT(id) as count
            FROM attendance
            WHERE class_id = ?
            GROUP BY student_id, status
        """, (class_id,)).fetchall()

        # Process data into a more usable format
        grades_by_student = {}
        for grade in grades:
            sid = grade['student_id']
            if sid not in grades_by_student:
                grades_by_student[sid] = []
            grades_by_student[sid].append({'subject': grade['subject_name'], 'score': grade['score']})

        attendance_by_student = {}
        for att in attendance_summary:
            sid = att['student_id']
            if sid not in attendance_by_student:
                attendance_by_student[sid] = {}
            attendance_by_student[sid][att['status']] = att['count']

        # 5. Compile the final report for each student
        student_reports = []
        for student in students:
            sid = student['id']
            student_grades = grades_by_student.get(sid, [])
            total_score = sum(g['score'] for g in student_grades if g['score'] is not None)
            num_subjects = len(student_grades)
            average = (total_score / num_subjects) if num_subjects > 0 else 0

            report = {
                'student_id': sid,
                'student_name': student['name'],
                'grades': student_grades,
                'attendance': {
                    'present': attendance_by_student.get(sid, {}).get('present', 0),
                    'absent': attendance_by_student.get(sid, {}).get('absent', 0),
                    'late': attendance_by_student.get(sid, {}).get('late', 0),
                },
                'total_score': total_score,
                'average': round(average, 2),
                'result': 'Pass' if average >= 50 else 'Fail'
            }
            student_reports.append(report)

        # 6. Calculate Rank
        student_reports.sort(key=lambda x: x['average'], reverse=True)
        for i, report in enumerate(student_reports):
            report['rank'] = i + 1

        final_data = {
            'class_info': dict(class_info),
            'student_results': student_reports
        }

        return jsonify(final_data)

    except Exception as e:
        print(f"---!!!! RESULTS REPORT ERROR !!!! --->: {e}")
        traceback.print_exc()
        return jsonify({'message': f'An internal error occurred: {e}'}), 500
    finally:
        if conn:
            conn.close()

# --- Student Report Card API Route ---
@app.route('/api/results/student-report/<int:student_id>', methods=['GET'])
@token_required
def get_student_report_card(student_id, **kwargs):
    exam_type = request.args.get('exam_type')
    if not exam_type:
        return jsonify({'message': 'Exam Type is required.'}), 400

    conn = None
    try:
        conn = get_db_connection()

        # 1. Get Student, Class, and Teacher Info
        student_info = conn.execute("""
            SELECT 
                s.name_km, s.name_en, s.name_jp, s.dob, s.photo_filename,
                c.name as class_name, c.academic_year,
                t.name as teacher_name
            FROM students s
            LEFT JOIN enrollments e ON s.id = e.student_id
            LEFT JOIN classes c ON e.class_id = c.id
            LEFT JOIN teachers t ON c.teacher_id = t.id
            WHERE s.id = ?
        """, (student_id,)).fetchone()

        if not student_info:
            return jsonify({'message': 'Student not found or not enrolled in a class.'}), 404

        class_id_row = conn.execute("SELECT class_id FROM enrollments WHERE student_id = ?", (student_id,)).fetchone()
        if not class_id_row:
             return jsonify({'message': 'Student is not enrolled in any class.'}), 404
        class_id = class_id_row['class_id']


        # 2. Get grades for the specified exam type
        grades = conn.execute("""
            SELECT s.name as subject_name, g.score
            FROM grades g
            JOIN subjects s ON g.subject_id = s.id
            WHERE g.student_id = ? AND g.exam_type = ? AND g.class_id = ?
        """, (student_id, exam_type, class_id)).fetchall()

        # 3. Get attendance summary
        attendance_summary = conn.execute("""
            SELECT status, COUNT(id) as count
            FROM attendance
            WHERE student_id = ? AND class_id = ?
            GROUP BY status
        """, (student_id, class_id)).fetchall()

        attendance_by_student = {att['status']: att['count'] for att in attendance_summary}

        # 4. Calculate total, average, and result
        total_score = sum(g['score'] for g in grades if g['score'] is not None)
        num_subjects = len(grades)
        average = (total_score / num_subjects) if num_subjects > 0 else 0
        
        # 5. Get rank within the class for this exam type
        all_class_averages = conn.execute("""
            SELECT student_id, AVG(score) as avg_score
            FROM grades
            WHERE class_id = ? AND exam_type = ?
            GROUP BY student_id
            ORDER BY avg_score DESC
        """, (class_id, exam_type)).fetchall()

        rank = -1
        for i, result in enumerate(all_class_averages):
            if result['student_id'] == student_id:
                rank = i + 1
                break
        
        final_report = {
            'student_info': dict(student_info),
            'grades': [dict(g) for g in grades],
            'attendance': {
                'present': attendance_by_student.get('present', 0),
                'absent': attendance_by_student.get('absent', 0),
                'late': attendance_by_student.get('late', 0),
            },
            'total_score': total_score,
            'average': round(average, 2),
            'rank': rank,
            'result': 'Pass' if average >= 50 else 'Fail'
        }
        
        return jsonify(final_report)

    except Exception as e:
        print(f"---!!!! STUDENT REPORT CARD ERROR !!!! --->: {e}")
        traceback.print_exc()
        return jsonify({'message': f'An internal error occurred: {e}'}), 500
    finally:
        if conn:
            conn.close()


# --- Export API Routes ---

# NEW HELPER FUNCTION FOR PDF IMAGE EMBEDDING
def image_to_base64_data_uri(filepath):
    """Reads an image file and converts it to a base64 data URI."""
    if not filepath or not os.path.exists(filepath):
        return None
    try:
        mime_type, _ = mimetypes.guess_type(filepath)
        if not mime_type or not mime_type.startswith('image'):
            return None
        with open(filepath, "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
        return f"data:{mime_type};base64,{encoded_string}"
    except Exception as e:
        print(f"Error converting image to base64: {e}")
        return None

@app.route('/api/students/export/excel')
@token_required
def export_students_excel(**kwargs):
    try:
        lang = request.args.get('lang', 'km') 
        
        headers_translations = {
            'km': ['ID', 'ឈ្មោះ (ខ្មែរ)', 'ឈ្មោះ (អង់គ្លេស)', 'ឈ្មោះ (ជប៉ុន)', 'ថ្ងៃខែឆ្នាំកំណើត', 'ទំនាក់ទំនង', 'អាសយដ្ឋាន', 'ឈ្មោះអាណាព្យាបាល', 'ទំនាក់ទំនងអាណាព្យាបាល'],
            'en': ['ID', 'Name (Khmer)', 'Name (English)', 'Name (Japanese)', 'Date of Birth', 'Contact', 'Address', 'Parent Name', 'Parent Contact'],
            'jp': ['ID', '名前 (クメール語)', '名前 (英語)', '名前 (日本語)', '生年月日', '連絡先', '住所', '保護者名', '保護者の連絡先']
        }

        conn = get_db_connection()
        students = conn.execute("SELECT id, name_km, name_en, name_jp, dob, contact, address, parent_name, parent_contact FROM students ORDER BY id DESC").fetchall()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Students"

        headers = headers_translations.get(lang, headers_translations['km'])
        ws.append(headers)

        if lang == 'km':
            font_name = "Khmer OS Battambang"
        elif lang == 'jp':
            font_name = "MS Gothic" 
        else: 
            font_name = "Arial"
            
        main_font = Font(name=font_name, size=11)
        bold_font = Font(name=font_name, size=12, bold=True)

        for cell in ws[1]:
            cell.font = bold_font

        for row in students:
            ws.append([
                row['id'], row['name_km'], row['name_en'], row['name_jp'], row['dob'], 
                row['contact'], row['address'] or '', 
                row['parent_name'] or '', row['parent_contact'] or ''
            ])

        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.font = main_font
        
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25


        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name="students_export.xlsx", as_attachment=True)
    except Exception as e:
        print(f"---!!!! EXCEL EXPORT ERROR !!!! --->: {e}")
        traceback.print_exc()
        return jsonify({'message': f'An error occurred during Excel export: {e}'}), 500


@app.route('/api/students/export/pdf')
@token_required
def export_students_pdf(**kwargs):
    try:
        lang = request.args.get('lang', 'km') 

        pdf_translations = {
            'km': { 'title': 'បញ្ជីឈ្មោះសិស្ស' },
            'en': { 'title': 'Student List' },
            'jp': { 'title': '学生一覧' }
        }
        
        t = pdf_translations.get(lang, pdf_translations['km'])

        conn = get_db_connection()
        students = conn.execute("""
            SELECT s.id, s.name_km, s.name_en, s.name_jp, s.dob, s.contact, s.address, s.photo_filename, c.name as class_name
            FROM students s
            LEFT JOIN enrollments e ON s.id = e.student_id
            LEFT JOIN classes c ON e.class_id = c.id
            GROUP BY s.id ORDER BY s.id DESC
        """).fetchall()
        conn.close()

        table_rows_html = ""
        for r in students:
            # --- START: PDF IMAGE FIX ---
            img_tag = "<div class='img-placeholder'></div>"
            if r['photo_filename']:
                photo_path = os.path.join(app.config['UPLOAD_FOLDER'], r['photo_filename'])
                data_uri = image_to_base64_data_uri(photo_path)
                if data_uri:
                    img_tag = f"<img src='{data_uri}'>"
            # --- END: PDF IMAGE FIX ---
            
            name_html = f"""
                <div class='name-km'>{r['name_km'] or ''}</div>
                <div class='name-en'>{r['name_en'] or ''}</div>
                <div class='name-jp'>{r['name_jp'] or ''}</div>
            """

            table_rows_html += f"""
                <tr>
                    <td>{img_tag}</td>
                    <td>{name_html}</td>
                    <td>{r['class_name'] or 'N/A'}</td>
                    <td>{r['address'] or ''}</td>
                    <td>{r['contact'] or ''}</td>
                </tr>
            """
        
        # --- START: PDF LOGO FIX ---
        logo_path = os.path.join(BASE_DIR, 'static', 'images', 'logo.png')
        logo_data_uri = image_to_base64_data_uri(logo_path)
        logo_tag = f"<img src='{logo_data_uri}' class='header-logo'>" if logo_data_uri else ""
        # --- END: PDF LOGO FIX ---

        css_string = """
        @font-face {{
            font-family: 'KhmerApp';
            src: url('file://{KHMER_TTF}') format('truetype');
        }}
        @font-face {{
            font-family: 'JapaneseApp';
            src: url('file://{JAPANESE_TTF}') format('truetype');
        }}
        * {{ font-family: 'KhmerApp', 'JapaneseApp', sans-serif; }}
        body {{ font-size: 10pt; }}
        .header {{
            display: flex; align-items: center; gap: 20px;
            padding-bottom: 15px; margin-bottom: 15px; border-bottom: 2px solid #000;
        }}
        .header-logo {{ width: 60px; height: 60px; }}
        .header-text h1 {{ margin: 0; font-size: 20pt; }}
        .header-text p {{ margin: 0; font-size: 12pt; color: #555; }}
        table {{ width: 100%; border-collapse: collapse; }}
        th, td {{
            border: 1px solid #ccc; padding: 8px;
            vertical-align: middle; text-align: left;
        }}
        th {{ background-color: #f2f2f2; font-weight: bold; text-align: center; }}
        td img, .img-placeholder {{
            width: 120px; height: 120px; 
            object-fit: cover; border-radius: 8px;
            display: block; margin: auto;
        }}
        .img-placeholder {{ background-color: #eee; }}
        .name-km {{ font-weight: bold; font-size: 1.1em; }}
        .name-en, .name-jp {{ font-size: 1em; color: #333; }}
        """.format(KHMER_TTF=KHMER_TTF, JAPANESE_TTF=JAPANESE_TTF)

        html_string = """
        <!DOCTYPE html>
        <html lang="{lang}">
        <head>
            <meta charset="utf-8">
            <title>{title}</title>
        </head>
        <body>
            <div class="header">
                {logo_tag}
                <div class="header-text">
                    <h1>YATAI School</h1>
                    <p>{title}</p>
                </div>
            </div>
            <table>
                <thead>
                    <tr>
                        <th style="width:20%">រូបថត / Photo / 写真</th>
                        <th style="width:30%">ឈ្មោះ / Name / 氏名</th>
                        <th style="width:15%">ថ្នាក់ / Class / クラス</th>
                        <th style="width:20%">អាសយដ្ឋាន / Address / 住所</th>
                        <th style="width:15%">ទំនាក់ទំនង / Contact / 連絡先</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows_html}
                </tbody>
            </table>
        </body>
        </html>
        """.format(
            lang=lang,
            title=t['title'],
            logo_tag=logo_tag,
            table_rows_html=table_rows_html
        )

        pdf_bytes = HTML(string=html_string, base_url=BASE_DIR).write_pdf(stylesheets=[CSS(string=css_string)])
        
        buf = io.BytesIO(pdf_bytes)
        buf.seek(0)
        return send_file(buf, download_name="student_list_report.pdf", as_attachment=True)

    except Exception as e:
        print(f"---!!!! PDF EXPORT ERROR !!!! --->: {e}")
        traceback.print_exc()
        return jsonify({'message': f'An error occurred during PDF export: {e}'}), 500

# --- Timetable API Routes ---
@app.route('/api/timetables', methods=['POST'])
@admin_required
def add_timetable_entry(**kwargs):
    data = request.get_json()
    class_id = data.get('class_id')
    teacher_id = data.get('teacher_id')
    subject_id = data.get('subject_id')
    day_of_week = data.get('day_of_week')
    start_time = data.get('start_time')
    end_time = data.get('end_time')

    if not all([class_id, teacher_id, subject_id, day_of_week, start_time, end_time]):
        return jsonify({'message': 'All fields are required.'}), 400

    conn = get_db_connection()
    try:
        conn.execute("""
            INSERT INTO timetables (class_id, teacher_id, subject_id, day_of_week, start_time, end_time)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (class_id, teacher_id, subject_id, day_of_week, start_time, end_time))
        conn.commit()
        return jsonify({'message': 'Timetable entry created successfully!'}), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'message': f'An error occurred: {e}'}), 500
    finally:
        if conn:
            conn.close()

@app.route('/api/timetables/class/<int:class_id>', methods=['GET'])
@token_required
def get_class_timetable(class_id, **kwargs):
    conn = get_db_connection()
    schedule = conn.execute("""
        SELECT tt.*, t.name as teacher_name, s.name as subject_name
        FROM timetables tt
        JOIN teachers t ON tt.teacher_id = t.id
        JOIN subjects s ON tt.subject_id = s.id
        WHERE tt.class_id = ?
        ORDER BY tt.day_of_week, tt.start_time
    """, (class_id,)).fetchall()
    conn.close()
    return jsonify([dict(row) for row in schedule])

@app.route('/api/timetables/<int:entry_id>', methods=['DELETE'])
@admin_required
def delete_timetable_entry(entry_id, **kwargs):
    conn = get_db_connection()
    conn.execute("DELETE FROM timetables WHERE id = ?", [entry_id])
    conn.commit()
    conn.close()
    return jsonify({'message': 'Timetable entry deleted successfully!'})

@app.route('/api/timetables/export/pdf')
@token_required
def export_timetable_pdf(**kwargs):
    class_id = request.args.get('class_id')
    lang = request.args.get('lang', 'km')
    if not class_id:
        return jsonify({'message': 'Class ID is required.'}), 400

    try:
        conn = get_db_connection()
        class_info = conn.execute("SELECT name FROM classes WHERE id = ?", (class_id,)).fetchone()
        schedule = conn.execute("""
            SELECT tt.*, t.name as teacher_name, s.name as subject_name
            FROM timetables tt
            JOIN teachers t ON tt.teacher_id = t.id
            JOIN subjects s ON tt.subject_id = s.id
            WHERE tt.class_id = ?
            ORDER BY tt.day_of_week, tt.start_time
        """, (class_id,)).fetchall()
        conn.close()

        if not class_info:
            return jsonify({'message': 'Class not found.'}), 404

        translations = {
            'km': {'title': 'កាលវិភាគសិក្សា', 'time': 'ម៉ោង', 'days': ['ចន្ទ', 'អង្គារ', 'ពុធ', 'ព្រហស្បតិ៍', 'សុក្រ', 'សៅរ៍', 'អាទិត្យ']},
            'en': {'title': 'Class Timetable', 'time': 'Time', 'days': ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']},
            'jp': {'title': 'クラスの時間割', 'time': '時間', 'days': ['月曜日', '火曜日', '水曜日', '木曜日', '金曜日', '土曜日', '日曜日']}
        }
        t = translations.get(lang, translations['km'])
        
        time_slots = [
            '07:00', '08:00', '09:00', '10:00', '11:00',
            '13:00', '14:00', '15:00', '16:00'
        ]

        # Build HTML grid
        table_header = f"<th>{t['time']}</th>"
        for day in t['days']:
            table_header += f"<th>{day}</th>"

        table_body = ""
        for time_slot in time_slots:
            table_body += f"<tr><td class='time-label'>{time_slot}</td>"
            for day_index in range(1, 8):
                entry_html = ""
                for entry in schedule:
                    if entry['day_of_week'] == day_index and entry['start_time'].startswith(time_slot):
                        entry_html += f"""
                            <div class='schedule-entry-pdf'>
                                <strong>{entry['subject_name']}</strong>
                                <p>{entry['teacher_name']}</p>
                            </div>
                        """
                table_body += f"<td>{entry_html}</td>"
            table_body += "</tr>"

        logo_path = os.path.join(BASE_DIR, 'static', 'images', 'logo.png')
        logo_data_uri = image_to_base64_data_uri(logo_path)
        logo_tag = f"<img src='{logo_data_uri}' class='header-logo'>" if logo_data_uri else ""
        
        css_string = """
        @font-face {{ font-family: 'KhmerApp'; src: url('file://{KHMER_TTF}'); }}
        @font-face {{ font-family: 'JapaneseApp'; src: url('file://{JAPANESE_TTF}'); }}
        * {{ font-family: 'KhmerApp', 'JapaneseApp', sans-serif; }}
        body {{ font-size: 9pt; }}
        .header {{ display: flex; align-items: center; gap: 20px; padding-bottom: 15px; margin-bottom: 15px; border-bottom: 2px solid #000; }}
        .header-logo {{ width: 60px; }}
        .header-text h1, .header-text p {{ margin: 0; }}
        table {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
        th, td {{ border: 1px solid #ccc; padding: 5px; vertical-align: top; text-align: center; height: 60px; }}
        th {{ background-color: #f2f2f2; font-weight: bold; }}
        .time-label {{ font-weight: bold; vertical-align: middle; }}
        .schedule-entry-pdf {{ background: #eef2ff; border-left: 3px solid #4f46e5; border-radius: 4px; padding: 4px; margin-bottom: 3px; text-align: left; font-size: 8pt; }}
        .schedule-entry-pdf strong {{ display: block; }}
        .schedule-entry-pdf p {{ margin: 2px 0 0; color: #555; }}
        """.format(KHMER_TTF=KHMER_TTF, JAPANESE_TTF=JAPANESE_TTF)

        html_string = """
        <!DOCTYPE html>
        <html>
        <head><title>Timetable</title></head>
        <body>
            <div class="header">
                {logo_tag}
                <div class="header-text">
                    <h1>{title}</h1>
                    <p>{class_name}</p>
                </div>
            </div>
            <table>
                <thead><tr>{table_header}</tr></thead>
                <tbody>{table_body}</tbody>
            </table>
        </body>
        </html>
        """.format(
            logo_tag=logo_tag,
            title=t['title'],
            class_name=class_info['name'],
            table_header=table_header,
            table_body=table_body
        )

        pdf_bytes = HTML(string=html_string, base_url=BASE_DIR).write_pdf(stylesheets=[CSS(string=css_string)])
        buf = io.BytesIO(pdf_bytes)
        buf.seek(0)
        return send_file(buf, download_name=f"timetable_{class_info['name']}.pdf", as_attachment=True)

    except Exception as e:
        print(f"---!!!! TIMETABLE PDF EXPORT ERROR !!!! --->: {e}")
        traceback.print_exc()
        return jsonify({'message': f'An error occurred during PDF export: {e}'}), 500


# --- Announcement API Routes ---
@app.route('/api/announcements', methods=['GET'])
@token_required
def get_announcements(**kwargs):
    conn = get_db_connection()
    announcements = conn.execute("""
        SELECT a.*, u.full_name as author_name
        FROM announcements a
        JOIN users u ON a.user_id = u.id
        ORDER BY a.created_at DESC
    """).fetchall()
    conn.close()
    return jsonify([dict(row) for row in announcements])

@app.route('/api/announcements', methods=['POST'])
@token_required
def add_announcement(**kwargs):
    data = request.get_json()
    title = data.get('title')
    content = data.get('content')
    user_id = kwargs['current_user']['id']

    if not title or not content:
        return jsonify({'message': 'Title and content are required.'}), 400

    conn = get_db_connection()
    try:
        conn.execute("INSERT INTO announcements (title, content, user_id) VALUES (?, ?, ?)",
                     (title, content, user_id))
        conn.commit()
        return jsonify({'message': 'Announcement created successfully!'}), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'message': f'An error occurred: {e}'}), 500
    finally:
        if conn:
            conn.close()

@app.route('/api/announcements/<int:announcement_id>', methods=['DELETE'])
@admin_required
def delete_announcement(announcement_id, **kwargs):
    conn = get_db_connection()
    conn.execute("DELETE FROM announcements WHERE id = ?", [announcement_id])
    conn.commit()
    conn.close()
    return jsonify({'message': 'Announcement deleted successfully!'})


# --- Run Application ---
if __name__ == '__main__':
    setup_database_and_admin()
    app.run(host='0.0.0.0', port=3000, debug=Fal